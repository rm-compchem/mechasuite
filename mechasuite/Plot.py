#!/usr/bin/python3
# -*- coding: utf-8 -*-

from PyQt5.QtWidgets import QMainWindow, QWidget, QApplication, QMenu, QComboBox, QInputDialog, QColorDialog, QLineEdit, QMessageBox, QFileDialog, QShortcut
from PyQt5.QtGui import QPainter, QPen, QFont, QColor, QPixmap, QBrush, QKeySequence
from PyQt5.QtCore import Qt, QRect
from mechasuite.Widgets import MultipleChoiceDialog
import sys
import numpy as np
import matplotlib
import os


class ComponentBase(object):
    def __init__(self, *args, **kwargs):
        if "color" in kwargs:
            self.color = kwargs["color"]
        else:
            self.color = None

        self.quad = []
        self.selected = False


class Tick(object):
    def __init__(self, **kwargs):
        if "n_ticks" in kwargs:
            self.n_ticks =kwargs["n_ticks"]
        else:
            self.n_ticks = 10

        self.labels = []
        self.label_offset = 40
        self.tcoors = []
        self.font = QFont("Arial", 10)
        self.length = 5
        self.color = QColor()

        if "lmax" in kwargs and "lmin" in kwargs:
            self.update_labels(kwargs["lmax"], kwargs["lmin"])

    def set_color(self, color):
        self.color = color

    def update_labels(self, ma, mi, n=0.5):
        # mi = int(mi)
        # while mi%n:
        #     mi+= 1
        # self.tcoors = [l for l in np.arange(mi, ma, n)]
        # self.labels = ["{: >4.0f}".format(l) for l in np.arange(mi, ma, n)]
        # esto de arriba es para tener intervalos redondos, pero no funciona muy bien
        self.tcoors = [l for l in np.linspace(mi, ma, 10)]
        self.labels = ["{: >4.1f}".format(l) for l in np.linspace(mi, ma, 10)]
        self.labels = np.append(self.labels, "0")
        self.tcoors = np.append(self.tcoors, 0.0)


class Label(ComponentBase):
    def __init__(self, parent, label, *args, **kwargs):
        ComponentBase.__init__(self, *args, **kwargs)
        self.parent = parent
        self.label = label
        self.name = label.name
        self.pixels = [0, 0]
        self.coors = label.coors
        self.text = label.text
        self.visible = label.visible
        self.color = QColor(label.color[0],
                            label.color[1],
                            label.color[2])
        if label.font == "":
            self.font = QFont()
        else:
            self.font = QFont()
            self.font.fromString(label.font)

        self.update_pixels()

    def hide(self):
        self.visible = False
        self.label.visible = False

    def show(self):
        self.visible = True
        self.label.visible = True

    def set_text(self, text):
        self.text = text
        self.label.text = text

    def set_family(self, family):
        self.font.setFamily(family)

    def set_point_size(self, s):
        self.font.setPointSize(s)
        self.label.font = self.font.toString()

    def set_color(self, color):
        # color has to be QColor instance
        self.color = color
        self.label.color = (color.red(),
                            color.green(),
                            color.blue())

    def get_color(self):
        return (self.color.red, self.color.green, self.color.blue)

    def update_coors(self):
        self.coors[0] = self.pixels[0] - self.parent.xmargin
        self.coors[0] /= (self.parent.width() - self.parent.xmargin * 2)
        self.coors[0] *= (self.parent.xmax - self.parent.xmin)

        self.coors[1] = self.pixels[1] - self.parent.ymargin
        self.coors[1] /= (self.parent.height() - self.parent.ymargin * 2)
        self.coors[1] *= (self.parent.ymax - self.parent.ymin)
        self.coors[1] = self.parent.ymax - self.coors[1]

    def update_pixels(self):
        self.pixels[0] = self.coors[0] / (self.parent.xmax - self.parent.xmin)
        self.pixels[0] *= (self.parent.width() - self.parent.xmargin * 2)
        self.pixels[0] += + self.parent.xmargin

        self.pixels[1] = (self.parent.ymax - self.coors[1])/ (self.parent.ymax - self.parent.ymin)
        self.pixels[1] *= (self.parent.height() - self.parent.ymargin * 2)
        self.pixels[1] += self.parent.ymargin

    def draw(self, qp):
        if not self.visible:
            return
        pen = QPen(self.color)
        qp.setPen(pen)
        # brush = QBrush()
        if self.selected:
            #brush.setColor(QColor(0,0,0))
            #qp.setBackgroundMode(Qt.OpaqueMode)
            #qp.setBackground(brush)
            self.font.setBold(True)
            self.font.setItalic(True)
        else:
            self.font.setBold(False)
            self.font.setItalic(False)
        qp.setFont(self.font)
        qp.drawText(int(self.pixels[0]), int(self.pixels[1]), self.text)

    def mouse_is_over(self, x, y):
        if self.pixels[0]-1 <= x <= self.pixels[0]+10 and \
               self.pixels[1]-10 <= y <= self.pixels[1] + 1:
            return True
        else:
            return False

    def translate_pixels(self, dxy):
        self.pixels[0] += dxy[0]
        self.pixels[1] += dxy[1]
        self.update_coors()


class Bar(ComponentBase):
    def __init__(self, parent, itm, *args, **kwargs):
        ComponentBase.__init__(self, *args, **kwargs)

        self.parent = parent
        self.coors = itm.coors
        self.itm = itm
        self.name = itm.name
        self.pixels = [0, 0, 0, 0]
        self.tp = "bar"
        self.blinks = []
        self.alinks = []
        self.color = QColor(itm.color[0],
                            itm.color[1],
                            itm.color[2])

        if "tp" in kwargs:
            self.tp = kwargs["tp"]

        if "width" in kwargs:
            self.width = kwargs["width"]
        else:
            self.width = itm.width

        self.clickable = []
        self.update_pixels()
        if self.tp == "bar":
            self.update_links()

        self.stdic = {
            "solid line": Qt.SolidLine,
            "dash line": Qt.DashLine,
            "dot line": Qt.DotLine,
            "dash dot line": Qt.DashDotLine,
            "dash dot dot line": Qt.DashDotDotLine
        }
        self.style = self.stdic[itm.style]

    def set_style(self, style):
        self.style = self.stdic[style]
        self.itm.style = style

    def set_color(self, color):
        # color has to be QColor instance
        self.color = color
        self.itm.color = (color.red(),
                            color.green(),
                            color.blue())

    def set_width(self, w):
        self.width = w
        self.itm.width = w

    def del_link(self, link):
        found = False
        for l in self.alinks:
            if l is link:
                self.alinks.remove(l)
                self.itm.del_link(link.name)
                found = True
                break
        if found:
            return
        for l in self.blinks:
            if l is link:
                self.blinks.remove(l)
                self.itm.del_link(link.name)

    def update_links(self):
        """Bar has to be instantiated after all the links are
        instantiated otherwise this is useless, not link object from
        plot will be fount"""

        for link in self.itm.blinks:
            for plink in self.parent.bars:
                if plink.tp == "link" and plink.name == link:
                    if plink not in self.blinks:
                        self.blinks.append(plink)

        for link in self.itm.alinks:
            for plink in self.parent.bars:
                if plink.tp == "link" and plink.name == link:
                    if plink not in self.alinks:
                        self.alinks.append(plink)

    def set_coors(self, coors):
        self.coors = coors
        self.update_pixels()

    def set_pixels(self, pixels):
        self.pixels = pixels
        self.update_coors()

    def translate_pixels(self, dxy):
        if self.tp == "link":
            return
        self.pixels[0] += dxy[0]
        self.pixels[2] += dxy[0]
        for b in self.blinks:
            b.set_pixel_x2(self.pixels[0])
        for b in self.alinks:
            b.set_pixel_x1(self.pixels[2])
        # self.pixels[1] += dxy[1]
        # self.pixels[3] += dxy[1]
        self.update_coors()

    def set_pixel_x1(self, x):
        self.pixels[0] = x
        self.coors[0] = self.pixels[0] - self.parent.xmargin
        self.coors[0] /= (self.parent.width() - self.parent.xmargin * 2)
        self.coors[0] *= (self.parent.xmax - self.parent.xmin)

    def set_pixel_x2(self, x):
        self.pixels[2] = x
        self.coors[2] = self.pixels[2] - self.parent.xmargin
        self.coors[2] /= (self.parent.width() - self.parent.xmargin * 2)
        self.coors[2] *= (self.parent.xmax - self.parent.xmin)

    def trans_point1_pixel(self, dx):
        self.pixels[0] += dx
        self.coors[0] = self.pixels[0] - self.parent.xmargin
        self.coors[0] /= (self.parent.width() - self.parent.xmargin * 2)
        self.coors[0] *= (self.parent.xmax - self.parent.xmin)

    def trans_point2_pixel(self, dx):
        self.pixels[2] += dx
        self.coors[2] = self.pixels[2] - self.parent.xmargin
        self.coors[2] /= (self.parent.width() - self.parent.xmargin * 2)
        self.coors[2] *= (self.parent.xmax - self.parent.xmin)

    def update_coors(self):
        self.coors[0] = self.pixels[0] - self.parent.xmargin
        self.coors[0] /= (self.parent.width() - self.parent.xmargin * 2)
        self.coors[0] *= (self.parent.xmax - self.parent.xmin)

        self.coors[2] = self.pixels[2] - self.parent.xmargin
        self.coors[2] /= (self.parent.width() - self.parent.xmargin * 2)
        self.coors[2] *= (self.parent.xmax - self.parent.xmin)

        self.coors[1] = self.pixels[1] - self.parent.ymargin
        self.coors[1] /= (self.parent.height() - self.parent.ymargin * 2)
        self.coors[1] *= (self.parent.ymax - self.parent.ymin)
        self.coors[1] = self.parent.ymax - self.coors[1]

        self.coors[3] = self.pixels[3] - self.parent.ymargin
        self.coors[3] /= (self.parent.height() - self.parent.ymargin * 2)
        self.coors[3] *= (self.parent.ymax - self.parent.ymin)
        self.coors[3] = self.parent.ymax - self.coors[3]

    def update_pixels(self):
        self.pixels[0] = self.coors[0] / (self.parent.xmax - self.parent.xmin)
        self.pixels[0] *= (self.parent.width() - self.parent.xmargin * 2)
        self.pixels[0] += + self.parent.xmargin

        self.pixels[2] = self.coors[2] / (self.parent.xmax - self.parent.xmin)
        self.pixels[2] *= (self.parent.width() - self.parent.xmargin * 2)
        self.pixels[2] += + self.parent.xmargin

        self.pixels[1] = (self.parent.ymax - self.coors[1])/ (self.parent.ymax - self.parent.ymin)
        self.pixels[1] *= (self.parent.height() - self.parent.ymargin * 2)
        self.pixels[1] += + self.parent.ymargin

        self.pixels[3] = (self.parent.ymax - self.coors[3]) / (self.parent.ymax - self.parent.ymin)
        self.pixels[3] *= (self.parent.height() - self.parent.ymargin * 2)
        self.pixels[3] += + self.parent.ymargin

    def mouse_is_over(self, x, y):
        m = (self.pixels[3]-self.pixels[1])/(self.pixels[2]-self.pixels[0])
        n = self.pixels[1]-m*self.pixels[0]
        if x < self.pixels[0] or x > self.pixels[2]:
            return False
        if m*x+n - (self.width + 5) <= y <= m*x+n + (self.width + 5):
            return True
        else:
            return False

    def draw(self, qp):
        pen = QPen(self.color, self.width, self.style)
        qp.setPen(pen)
        qp.drawLine(int(self.pixels[0]), int(self.pixels[1]), int(self.pixels[2]), int(self.pixels[3]))
        if self.selected:
            pen = QPen(Qt.blue, 2, Qt.DotLine)
            qp.setPen(pen)
            qp.drawEllipse(int(self.pixels[0] - self.width*1.5/2.), int(self.pixels[1] - self.width*1.5/2.), int(self.width*1.5), int(self.width*1.5))
            qp.drawEllipse(int(self.pixels[2] - self.width*1.5/2.), int(self.pixels[3] - self.width*1.5/2.), int(self.width*1.5), int(self.width*1.5))

    def delete(self, qp):
        # qp.eraseRect(self.pixels[0]-2, self.pixels[1]-2, self.pixels[2]+2, 4)
        r = QRect(0,0, self.parent.width(), self.parent.height())
        qp.eraseRect(r)


class Axis(ComponentBase):
    def __init__(self, parent, *args, **kwargs):
        ComponentBase.__init__(self, *args, **kwargs)

        self.parent = parent
        self.ori = [0, 0]
        self.xmax = 0
        self.ymax = 0
        self.update_max()

        if "xlabel" in kwargs:
            self.xlabel = kwargs["xlabel"]
        else:
            self.xlabel = "xlabel"

        if "ylabel" in kwargs:
            self.ylabel = kwargs["ylabel"]
        else:
            self.ylabel = "Relative Energy"

        if "width" in kwargs:
            self.width = kwargs["width"]
        else:
            self.width = 2

        self.yticks = Tick(pmax=self.ymax, pmin=self.ori[1])

        self.font = QFont("Arial", 10)

    # def set_

    def update_max(self):
        self.ori = [self.parent.xmargin, self.parent.height()-self.parent.ymargin]
        self.xmax = self.parent.width()-self.parent.xmargin
        self.ymax = self.parent.ymargin

    def draw(self, qp):
        self.update_max()
        pen = QPen(Qt.black, self.width, Qt.SolidLine)
        qp.setPen(pen)
        # draw x axis
        qp.drawLine(int(self.ori[0]), int(self.ori[1]), int(self.xmax), int(self.ori[1]))
        # draw y axis
        qp.drawLine(int(self.ori[0]), int(self.ori[1]), int(self.ori[0]), int(self.ymax))
        # draw x arrows
        qp.drawLine(int(self.parent.width() - self.parent.xmargin),
                    int(self.parent.height() - self.parent.ymargin),
                    int(self.parent.width() - self.parent.xmargin-8),
                    int(self.parent.height() - self.parent.ymargin-5))
        qp.drawLine(int(self.parent.width() - self.parent.xmargin),
                    int(self.parent.height() - self.parent.ymargin),
                    int(self.parent.width() - self.parent.xmargin-8),
                    int(self.parent.height() - self.parent.ymargin+5))

        qp.setFont(QFont("Arial", 12))

        self.yticks.update_labels(self.parent.ymax, self.parent.ymin)
        qp.setFont(self.yticks.font)
        # for t, l in zip(self.yticks.tcoors, self.yticks.labels):
        #     qp.drawText(self.ori[0]-self.yticks.label_offset, t, l)
        #     qp.drawLine(self.ori[0], t, self.ori[0]-self.yticks.length, t)
        ticks = self.parent.ycoor_to_ypixel(self.yticks.tcoors)
        #print(ticks)
        #print(self.yticks.tcoors)
        for l, t in zip(self.yticks.labels, ticks):
            qp.drawText(int(self.ori[0]-self.yticks.label_offset), int(t), l)
            qp.drawLine(int(self.ori[0]), int(t), int(self.ori[0]-self.yticks.length), int(t))


class Plot(QWidget):
    def __init__(self, parent, gd):
        QWidget.__init__(self, parent)
        # dictionary with the grapths data
        self.gd = gd
        # self.parent = parent
        # self.current_plot =
        self.qp = QPainter()
        self.xmargin_ratio, self.ymargin_ratio = 0.08, 0.05
        self.xmargin, self.ymargin = 50, 5
        self.xmin, self.xmax = 0, 10
        self.ymin, self.ymax = -10, 0
        self.axis = Axis(self)
        self.bars = []
        self.labels = []
        self.selected = []
        self.old_mouse_coors = (0, 0)
        self.delta_mouse_coors = (0, 0)
        self.left_down = False
        self.right_down = False
        self.middle_down = False
        self.stdic = {
            "solid line": Qt.SolidLine,
            "dash line": Qt.DashLine,
            "dot line": Qt.DotLine,
            "dash dot line": Qt.DashDotLine,
            "dash dot dot line": Qt.DashDotDotLine
        }
        self.rev_stdic = {
            Qt.SolidLine: "solid line",
            Qt.DashLine: "dash line",
            Qt.DotLine: "dot line",
            Qt.DashDotLine: "dash dot line",
            Qt.DashDotDotLine: "dash dot dot line"
        }

        self.palete = self.palette()
        self.palete.setColor(self.backgroundRole(), Qt.white)
        self.setPalette(self.palete)

        self.setMouseTracking(True)

        self.update_data()

    def update_data(self, gname=None):
        # this function updates only the graphics elements
        # which links are drawn
        # MAYBE I CAN CALL THE UPDATE RELATIVE ENERGIES OF LINKSKKKK that is coors of Data PLot elemetnts
        if gname is None:
            g = self.gd.get_plots()
            if len(g) > 0:
                g = g[0]
            else:
                g = None
        else:
            g = self.gd.get_plot(gname)
        if g is None:
            return
        self.current_plot = g
        #g.update_type_en()

        self.bars, self.labels = [], []
        for link in g.get_links():
            self.bars.append(Bar(self, link, tp="link"))
        for itm in g.get_itms():
            self.bars.append(Bar(self, itm))
        for label in g.get_labels():
            self.labels.append(Label(self, label))
        self.xmax = g.xmax
        self.ymax = g.ymax
        self.xmin = g.xmin
        self.ymin = g.ymin
        self.update()

    def contextMenuEvent(self, event):
        menu = QMenu(self)
        selectmenu = menu.addMenu("Select")
        barwith = menu.addAction("Bar Width")
        barstyle = menu.addAction("Bar Style")
        changecolor = menu.addAction("Change Color")
        itm_connect = menu.addAction("Connect")
        export_img = menu.addAction("Export PNG")
        export_react_net = menu.addAction("Export Reaction Network")
        export_graph_to_excel = menu.addAction("Export Plot to XLSX")
        hide_labels = menu.addAction("Hide Labels")
        show_labels = menu.addAction("Show Labels")
        clone_graph = menu.addAction("Clone Plot")
        showmatplotlib = menu.addAction("Show With Matplotlib")
        show_mult_matplotlib = menu.addAction("Show Multiple Plots With Matplotlib")

        select_bars = selectmenu.addAction("Select All Bars")
        select_links = selectmenu.addAction("Select All Links")
        select_labels = selectmenu.addAction("Select All Labels")

        action = menu.exec_(self.mapToGlobal(event.pos()))
        if action == barwith:
            self.on_bar_with()
        elif action == barstyle:
            self.on_bar_style()
        elif action == changecolor:
            self.on_change_color()
        elif action == itm_connect:
            self.itm_connect()
        elif action == select_bars:
            self.select_bars()
        elif action == select_links:
            self.select_links()
        elif action == select_labels:
            self.select_labels()
        elif action == export_img:
            self.save_img()
        elif action == export_react_net:
            self.export_reaction_network()
        elif action == hide_labels:
            self.hide_labels()
        elif action == show_labels:
            self.show_labels()
        elif action == export_graph_to_excel:
            self.on_export_graph_to_excel()
        elif action == clone_graph:
            self.on_clone_graph()
        elif action == showmatplotlib:
            self.show_matplotlib()
        elif action == show_mult_matplotlib:
            self.show_multiple_matplotlib()

    def plot_current_in_matplotlib(self, plt):
        stdic = {
            "solid line": "-",
            "dash line": "--",
            "dot line": "-.",
            "dash dot line": ":",
            "dash dot dot line": ":"
        }

        for bar in self.bars:
            color = '#%02x%02x%02x' % (bar.color.red(), bar.color.green(), bar.color.blue())
            plt.plot([bar.coors[0], bar.coors[2]], [bar.coors[1], bar.coors[3]],
                     stdic[self.rev_stdic[bar.style]], color=color,
                     linewidth=bar.width, solid_capstyle="projecting", fillstyle="none")
        #plt.xlim(self.xmin, self.xmax)
        #plt.ylim(self.ymin, self.ymax)

        for label in self.labels:
            if label.visible:
                plt.text(label.coors[0], label.coors[1], label.text, fontsize=label.font.pointSize())

    def setup_maplotlib_settings(self, plt, xmax, xmin, ymax, ymin):
        print("setting matplotlib ", xmax, xmin, ymax, ymin)
        fontsize = int(self.parent().comboSize.currentText())

        # matplotlib.rcParams("")
        matplotlib.rc('xtick', labelsize=fontsize * 0.8)
        matplotlib.rc('ytick', labelsize=fontsize * 0.8)
        matplotlib.rcParams["savefig.directory"] = os.getcwd()
        # matplotlib.rc('text', usetex=True)
        # matplotlib.rcParams['text.latex.preamble'] = [r"\usepackage{amsmath}"]

        matplotlib.rcParams["font.family"] = "Times"
        matplotlib.rcParams["figure.figsize"] = (8,4)
        plt.tight_layout()
        #plt.subplots_adjust(left=0.05, bottom=0.05, top=0.95, right=0.95, wspace=0.2, hspace=0.26)

        plt.xlim(xmin, xmax)
        plt.ylim(ymin, ymax)

        plt.xlabel(r"$Reaction\ Coordinate$", fontsize=fontsize)
        # plt.xlabel("Reaction Coordinate", fontsize=fontsize, fontweight="bold")
        if self.current_plot.etype == "G":
            plt.ylabel(r"$G_{rel}\ (kcal/mol)$", fontsize=fontsize, fontweight="bold")
        else:
            # plt.ylabel("$\boldsymbol{E_{rel} (kcal/mol)}$", fontsize=fontsize, fontweight="bold")
            plt.ylabel(r"${E_{rel}\ (kcal/mol)}$", fontsize=fontsize, fontweight="bold")
        plt.xticks([])
        return plt

    def plot_graph_in_matplotlib(self, plt, gname):
        g = self.gd.get_plot(gname)
        bars = []; labels = []

        for link in g.get_links():
            bars.append(Bar(self, link, tp="link"))
        for itm in g.get_itms():
            bars.append(Bar(self, itm))
        for label in g.get_labels():
            labels.append(Label(self, label))

        stdic = {
            "solid line": "-",
            "dash line": "--",
            "dot line": "-.",
            "dash dot line": ":",
            "dash dot dot line": ":"
        }

        for bar in bars:
            color = '#%02x%02x%02x' % (bar.color.red(), bar.color.green(), bar.color.blue())
            plt.plot([bar.coors[0], bar.coors[2]], [bar.coors[1], bar.coors[3]],
                     stdic[self.rev_stdic[bar.style]], color=color,
                     linewidth=bar.width, solid_capstyle="projecting", fillstyle="none")

        for label in labels:
            if label.visible:
                plt.text(label.coors[0], label.coors[1], label.text, fontsize=label.font.pointSize())

    def show_multiple_matplotlib(self):
        g = self.gd.get_plots()
        gnames = [i.name for i in g]
        dialog = MultipleChoiceDialog(gnames)
        dialog.exec()
        if not dialog.ok:
            return

        from matplotlib import pyplot as plt

        # global max min
        xmax = -1e6; xmin=1e6; ymax=-1e6; ymin=1e6
        for graph_name in dialog.items:
            gi = self.gd.get_plot(graph_name)
            if xmax < gi.xmax: xmax = gi.xmax
            if xmin > gi.xmin: xmin = gi.xmin
            if ymax < gi.ymax: ymax = gi.ymax
            if ymin > gi.ymin: ymin = gi.ymin

        for graph_name in dialog.items:
            #self.update_data(graph_name)
            #self.plot_current_in_matplotlib(plt)
            self.plot_graph_in_matplotlib(plt, graph_name)

        plt = self.setup_maplotlib_settings(plt, xmax, xmin, ymax, ymin)

        plt.show()

    def show_matplotlib(self):
        from matplotlib import pyplot as plt

        self.setup_maplotlib_settings(plt, self.xmax, self.xmin, self.ymax, self.ymin)
        self.plot_current_in_matplotlib(plt)

        plt.show()

    def on_clone_graph(self):
        newname,ok = QInputDialog.getText(self, "New name", "Enter New name")
        if not ok or not newname:
            return
        self.gd.clone_plot(self.current_plot.name, newname)
        #print("plot names: ", self.gd.get_plots_names())
        #self.parent().gd = self.gd
        #self.update_data(self.current_plot.name)
        self.parent().update_data()

    def on_export_graph_to_excel(self):
        try:
            import openpyxl
            from openpyxl.drawing.text import CharacterProperties, ParagraphProperties, Font
            from openpyxl.chart import (ScatterChart,
                                    Reference,
                                    Series,)
        except:
            diag = QMessageBox(self)
            diag.setText("openpyxl is not installed")
            diag.exec()
            return
        
        chart = ScatterChart()
        chart.title = None
        chart.style = 14

        font = Font(typeface='Verdana')
        size = 1200  # 14 point size
        cp = CharacterProperties(latin=font, sz=size, b=False)  # Not bold
        pp = ParagraphProperties(defRPr=cp)
        chart.x_axis.title = 'Reaction Coordinate'
        chart.x_axis.title.tx.rich.p[0].pPr = pp
        chart.y_axis.title = 'Relative Energy'
        chart.y_axis.title.tx.rich.p[0].pPr = pp
        chart.y_axis.majorGridlines = None
        chart.y_axis.minorGridlines = None
        chart.x_axis.majorGridlines = None
        chart.x_axis.minorGridlines = None
        chart.y_axis.scaling.max = round(self.ymax)
        chart.y_axis.scaling.min = round(self.ymin)
        chart.x_axis.scaling.max = round(self.xmax)
        chart.x_axis.scaling.min = round(self.xmin)
        chart.x_axis.tickLblPos = "low"
        chart.legend = None
        stdic = {
            "solid line": "solid",
            "dash line": "dash",
            "dot line": "dot",
            "dash dot line": "dashDot",
            "dash dot dot line": "sysDashDotDot"
        }
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["", ""])
        row = 2
        for bar in self.bars:
            # if bar.tp == "link":
            #     continue
            _ = ws.cell(column=1, row=row, value=bar.coors[0])
            _ = ws.cell(column=2, row=row, value=bar.coors[1])
            #ws.append(bar.coors[:2])
            row += 1
            _ = ws.cell(column=1, row=row, value=bar.coors[2])
            _ = ws.cell(column=2, row=row, value=bar.coors[3])
            #ws.append(bar.coors[2:])
            row += 1
            xvalues = Reference(ws, min_col=1, min_row=row-2, max_row=row-1)
            values = Reference(ws, min_col=2, min_row=row-2, max_row=row-1)
            series = Series(values, xvalues, title_from_data=False)
            color = '%02x%02x%02x' % (bar.color.red(), bar.color.green(), bar.color.blue())
            series.graphicalProperties.line.solidFill = color
            series.graphicalProperties.line.dashStyle = stdic[self.rev_stdic[bar.style]]
            series.graphicalProperties.line.width = bar.width*13000 # width in EMUs
            # if bar.tp == "bar":
            #     series.graphicalProperties.line.cap = "flat" # width in EMUs
            chart.series.append(series)
        ws.add_chart(chart, "C2")
        wb.save("newplot.xlsx")

    def hide_labels(self):
        for i in self.selected:
            if isinstance(i, Label):
                i.hide()
        self.update()

    def show_labels(self):
        for i in self.selected:
            if isinstance(i, Label):
                i.show()
        self.update()

    def export_reaction_network(self):
        # if plott type is G
        temps = [str(i) for i in self.current_plot.temps]
        dialog = MultipleChoiceDialog(temps)
        dialog.exec()
        if not dialog.ok:
            return
        temps = [float(i) for i in dialog.items]
        # considre also plot type E
        # .... bla bla

        filename = QFileDialog.getSaveFileName()
        if not filename[0]:
            return
        self.current_plot.export_reaction_network(filename[0])

    def select_bars(self):
        self.selected = []
        for bar in self.bars:
            if bar.tp == "bar":
                self.selected.append(bar)
                bar.selected = True
        self.update()

    def select_links(self):
        self.selected = []
        for bar in self.bars:
            if bar.tp == "link":
                self.selected.append(bar)
                bar.selected = True
        self.update()

    def select_labels(self):
        self.selected = []
        for label in self.labels:
            self.selected.append(label)
            label.selected = True
        self.update()

    def itm_connect(self):
        sortable = [itm for itm in self.selected
                     if isinstance(itm, Bar) and itm.tp == "bar"]

        allxs = [itm.coors[0] for itm in sortable]
        allxs.sort()
        sorted_itms = [itm for x in allxs for itm in sortable
                       if itm.coors[0] == x]

        for n, itm in enumerate(sorted_itms):
            if n == len(sorted_itms) - 1:
                break
            if self.bars_connected(itm, sorted_itms[n+1]):
                continue
            pitm1 = itm.itm
            pitm2 = sorted_itms[n+1].itm
            newlink = self.current_plot.add_link(pitm1, pitm2)
            if newlink is None:
                continue
            self.bars.append(Bar(self, newlink, tp="link"))
            itm.update_links()
            sorted_itms[n+1].update_links()
        self.update()

    def on_change_color(self):
        color = QColorDialog().getColor()
        for b in self.selected:
            b.set_color(color)

    def on_bar_style(self):
        dialog = MultipleChoiceDialog(list(self.stdic.keys()))
        dialog.exec()
        if not dialog.ok:
            return
        style = dialog.items[0]
        for b in self.selected:
            if isinstance(b, Bar):
                b.set_style(style)

    def on_bar_with(self):
        if not self.selected:
            return
        w, ok = QInputDialog().getText(self, "Enter Width", "Width")
        if not ok:
            return
        try:
            w = int(w)
        except ValueError:
            return

        for b in self.selected:
            if isinstance(b, Bar):
                b.set_width(w)
        self.update()

    def paintEvent(self, e):
        self.qp.begin(self)
        self.qp.setRenderHint(QPainter.Antialiasing)
        self.update_margin()
        # self.drawLines(self.qp)
        self.axis.draw(self.qp)
        for bar in self.bars:
            bar.update_pixels()
            bar.draw(self.qp)
        for label in self.labels:
            label.update_pixels()
            label.draw(self.qp)
        self.qp.end()

    def save_img(self):
        scale = 1.1
        w, h = int(self.width()*scale), int(self.height()*scale)
        pixmap = QPixmap(w, h)
        p  = Plot(None, self.gd)
        p.resize(w, h)
        p.update_data()
        p.render(pixmap)
        pixmap.save("salida.png")

    def bars_connected(self, bar1, bar2):
        for link1 in bar1.alinks:
            for link2 in bar2.blinks:
                if link1.name == link2.name:
                    return True
        for link1 in bar1.blinks:
            for link2 in bar2.alinks:
                if link1.name == link2.name:
                    return True
        return False

    def coor_to_pixel(self, xcoor, ycoor):
        x = xcoor/(self.xmax - self.xmin)
        x *= (self.width()-self.xmargin*2)
        x += + self.xmargin
        y = (self.ymax - ycoor)/(self.ymax - self.ymin)
        y *= (self.height()-self.ymargin*2)
        y += self.ymargin
        return x, y

    def ycoor_to_ypixel(self, ycoors):
        pixelcoors = []
        if self.ymax - self.ymin == 0.0:
            self.ymax += 0.01
            self.ymin -= 0.01

        for yc in ycoors:
            y = (self.ymax - yc) / (self.ymax - self.ymin)
            y *= (self.height() - self.ymargin * 2)
            y += self.ymargin
            pixelcoors.append(y)
        return pixelcoors

    def update_margin(self):
        self.xmargin = self.xmargin_ratio*self.width()
        self.ymargin = self.ymargin_ratio*self.width()

    def get_clicked(self, x, y):
        obj = None
        for b in self.bars:
            if b.mouse_is_over(x, y):
                obj = b

        for b in self.labels:
            if b.mouse_is_over(x, y):
                obj = b
        return obj

    def unselect(self, objs = []):
        if objs:
            for b in objs:
                b.selected = False
                try:
                    self.selected.remove(b)
                except ValueError:
                    pass
        else:
            self.selected = []
            for b in self.bars:
                b.selected = False

            for b in self.labels:
                b.selected = False

    def on_key_del(self):
        for itm in self.selected:
            if isinstance(itm, Bar):
                self.del_bar(itm)
            elif isinstance(itm, Label):
                self.labels.remove(itm)
                del self.current_plot.labels[itm.name]
        self.selected = []
        self.update()

    def del_link(self, link):
        try:
            self.bars.remove(link)
        except ValueError:
            pass
        else:
            del self.current_plot.links[link.name]
        for bar in self.bars:
            bar.del_link(link)

    def del_bar(self, bar):
        if bar.tp == "bar":
            for link in bar.alinks:
                self.del_link(link)
            for link in bar.blinks:
                self.del_link(link)
            self.bars.remove(bar)
            del self.current_plot.itms[bar.name]

        elif bar.tp == "link":
            self.del_link(bar)

    # HERE ARE THE EVNET HANDLERS---------------------------
    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Escape:
            self.unselect()
            self.update()
            # self.parentWidget().close()
        if e.key() == Qt.Key_A:
            self.selected = []
            for bar in self.bars:
                self.selected.append(bar)
                bar.selected = True
            self.update()
        if e.key() == Qt.Key_Delete:
            self.on_key_del()
        if e.key() == Qt.Key_B:
            for i in self.selected:
                if isinstance(i, Label):
                    i.set_family("Arial")
        if e.key() == Qt.Key_K and e.modifiers() == Qt.ControlModifier:
            for itm in self.bars: # bien burro esto
                if itm.tp != "bar": continue
                for label in self.labels:
                    if not label.selected: continue
                    if label.label.itm.name == itm.itm.name:
                        label.coors = itm.itm.coors[:]
                        break
            self.update()

    def wheelEvent(self, e):
        deltay = (self.ymax-self.ymin)/10
        if e.angleDelta().y() > 0:
            # self.xmax += 1
            # self.xmin -= 1
            self.ymax += deltay
            self.ymin -= deltay
        else:
            # self.xmax -= 1
            # self.xmin += 1
            self.ymax -= deltay
            self.ymin += deltay
        self.update()

    def mouseMoveEvent(self, e):
        self.unsetCursor()
        x, y = e.x(), e.y()
        self.delta_mouse_coors = (x - self.old_mouse_coors[0], y - self.old_mouse_coors[1])
        self.old_mouse_coors = (x, y)
        if self.left_down:
            for itm in self.selected:
                itm.translate_pixels(self.delta_mouse_coors)
        if self.middle_down:
            deltay = (self.ymax - self.ymin) / 400
            deltax = (self.xmax - self.xmin) / 400
            self.ymax += deltay*self.delta_mouse_coors[1]
            self.ymin += deltay*self.delta_mouse_coors[1]
            self.xmax += deltax*self.delta_mouse_coors[0]
            self.xmin += deltax*self.delta_mouse_coors[0]
            # print(self.xmax)

        for b in self.bars:
            if b.mouse_is_over(x, y):
                self.setCursor(Qt.SizeAllCursor)
        for b in self.labels:
            if b.mouse_is_over(x, y):
                self.setCursor(Qt.SizeAllCursor)
        self.update()

    def mousePressEvent(self, e):
        self.setFocus()
        if e.button() == 4: self.middle_down = True
        x, y = e.x(), e.y()
        if e.button() == 1:
            self.left_down = True
            if e.modifiers() == Qt.ControlModifier:
                clicked = self.get_clicked(x, y)
                if clicked is not None:
                    if clicked in self.selected:
                        self.unselect([clicked])
                    else:
                        self.selected.append(clicked)
                        clicked.selected = True
            elif e.modifiers() == Qt.ShiftModifier:
                clicked = self.get_clicked(x, y)
                if isinstance(clicked, Bar):
                    self.select_connected_bars(clicked, [clicked])
            elif e.modifiers() == Qt.AltModifier:
                clicked = self.get_clicked(x, y)
                if isinstance(clicked, Bar):
                    self.select_connected_bars(clicked, [clicked], onlybar=True)
            else:
                if self.selected: # confusing, this is a list instead of bool
                    clicked = self.get_clicked(x, y)
                    if clicked is None:
                        self.unselect()
                    else:
                        if clicked not in self.selected:
                            self.unselect()
                            self.selected = [clicked]
                            clicked.selected = True
                else:
                    clicked = self.get_clicked(x, y)
                    if clicked is not None:
                        self.selected.append(clicked)
                        # print (clicked.pixels)
                        clicked.selected = True
        elif e.button() == 2:
            self.right_down = True
        elif e.button() == 4:
            self.middle_down = True

    def mouseDoubleClickEvent(self, e):
        self.unselect()
        clicked = self.get_clicked(e.x(), e.y())
        if isinstance(clicked, Label):
            newtext, ok = QInputDialog().getText(self, "Enter Text", "Text", QLineEdit.Normal,  clicked.text)
            if not ok:
                return
            clicked.set_text(newtext)

    def mouseReleaseEvent(self, e):
        if e.button() == 1: self.left_down = False
        elif e.button() == 2: self.right_down = False
        elif e.button() == 4: self.middle_down = False
        self.update()

    def on_parentcombo_size(self, size):
        for b in self.selected:
            if isinstance(b, Label):
                b.set_point_size(int(size))
        self.update()

    def on_parentcomboE(self, t):
        self.current_plot.update_type_en(t)
        if self.current_plot.etype == "G":
            self.parent().comboT.clear()
            self.parent().comboT.addItems([str(t)
                                           for t in self.current_plot.temps])
            self.parent().comboT.setCurrentIndex(0)
            self.parent().comboT.update()
        else:
            self.parent().comboT.clear()
        self.update_data(self.current_plot.name)

    def on_parentcombo_temp(self, t):
        self.current_plot.update_temp_en(t)
        self.update_data(self.current_plot.name)

    def select_connected_bars(self, baritm, checked, onlybar=False):
        self.selected.append(baritm)
        checked.append(baritm)
        for itm in baritm.blinks:
            if not onlybar:
                itm.selected = True
                self.selected.append(itm)
            checked.append(itm)
            for bar in self.bars:
                if bar in checked: continue
                if itm in bar.blinks or itm in bar.alinks:
                    checked.append(bar)
                    self.select_connected_bars(bar, checked, onlybar)
        for itm in baritm.alinks:
            if not onlybar:
                itm.selected = True
                self.selected.append(itm)
            checked.append(itm)
            for bar in self.bars:
                if bar in checked: continue
                if itm in bar.blinks or itm in bar.alinks:
                    checked.append(bar)
                    self.select_connected_bars(bar, checked, onlybar)
        self.update()
        

class MainCanvas(QMainWindow):
    def __init__(self, parent, gd):
        QMainWindow.__init__(self, parent)
        # TOOL BAR ----------------------------
        self.toolbar = self.addToolBar("toolbar")

        self.comboG = QComboBox(self.toolbar)
        self.comboSize = QComboBox(self.toolbar)
        self.comboE = QComboBox(self.toolbar)
        self.comboT = QComboBox(self.toolbar)

        self.data = gd
        self.init_ui()
        self.plot = Plot(self, gd)
        self.setCentralWidget(self.plot)
        self.bindings()

    def bindings(self):
        self.comboG.currentTextChanged.connect(self.on_combog)
        self.comboSize.currentTextChanged.connect(self.on_combo_size)
        self.comboE.currentTextChanged.connect(self.on_comboE)
        self.comboT.currentIndexChanged.connect(self.on_combo_temp)

    def init_ui(self):
        self.comboG.setEditable(True)
        self.comboG.addItems(self.data.get_plots_names())
        self.comboSize.addItems([str(i) for i in range(30)])
        self.comboSize.setCurrentIndex(12)

        self.comboE.addItems(["E", "G"])
        self.comboT.addItems([" "])
        self.comboT.setSizeAdjustPolicy(0)

        self.toolbar.addWidget(self.comboG)
        self.toolbar.addWidget(self.comboSize)
        self.toolbar.addWidget(self.comboE)
        self.toolbar.addWidget(self.comboT)

        self.setGeometry(200, 200, 800, 800)
        self.setWindowTitle('Plot')
        self.palete = self.palette()
        self.palete.setColor(self.backgroundRole(), Qt.white)
        self.setPalette(self.palete)

        quit_shortcut = QShortcut(QKeySequence('Ctrl+Q'), self)
        quit_shortcut.activated.connect(self.close)
        # self.show()

    def update_data(self):
        text = self.comboG.currentText()
        gnames = self.data.get_plots_names()
        self.comboG.clear()
        self.comboG.addItems(gnames)
        if text in gnames:
            gname = text # graph name
            self.comboG.setCurrentText(text)
        elif len(gnames) > 0:
            gname = gnames[0]
            self.comboG.setCurrentText(gnames[0])
        else:
            gname = ""
            self.comboG.setCurrentText("")
        
        # update the qt graphics plot
        self.plot.on_parentcomboE(self.comboE.currentText())
        self.plot.update_data(gname)

    def on_combog(self):
        self.plot.update_data(self.comboG.currentText())

    def on_combo_size(self):
        self.plot.on_parentcombo_size(self.comboSize.currentText())

    def on_comboE(self):
        self.plot.on_parentcomboE(self.comboE.currentText())

    def on_combo_temp(self):
        try:
            temp = float(self.comboT.currentText())
        except ValueError:
            return
        else:
            self.plot.on_parentcombo_temp(temp)


def plot(graph_data):
    # app = QApplication(sys.argv)
    ex = MainCanvas(graph_data)
    # app.exec_()
    return ex


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MainCanvas()
    sys.exit(app.exec_())
